from __future__ import annotations


import calendar
import math
import re
from dataclasses import dataclass
from datetime import datetime, date, time, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple


import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter


try:
    import xlrd
except Exception:
    xlrd = None


# ── Date Parsing Patterns ─────────────────────────────────────────────────────
DATE_PATTERNS = [
    "%d/%m/%Y %I:%M:%S %p",
    "%d/%m/%Y %I:%M %p",
    "%m/%d/%Y %I:%M:%S %p",
    "%m/%d/%Y %I:%M %p",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
]


# ── Default Employee Config ───────────────────────────────────────────────────
DEFAULT_EMPLOYEE_CONFIG: Dict[int, Dict[str, Any]] = {
    4:  {"short_name": "Cris",    "full_name": "Baliton, Maria Cristina E.",        "schedule_in": "09:00", "schedule_out": "17:00"},
    7:  {"short_name": "Kazz",    "full_name": "Pavia, Kazzela Aira C.",            "schedule_in": "09:00", "schedule_out": "18:00"},
    8:  {"short_name": "Janneth", "full_name": "Duran, Janneth B.",                 "schedule_in": "06:00", "schedule_out": "15:00"},
    9:  {"short_name": "Mark",    "full_name": "Vibat, Mark Kevin",                 "schedule_in": "09:00", "schedule_out": "18:00"},
    11: {"short_name": "Yvonne",  "full_name": "Badua, Yvonney B.",                 "schedule_in": "06:00", "schedule_out": "15:00"},
    13: {"short_name": "Charles", "full_name": "Marvida, Charles Andrew P.",        "schedule_in": "06:00", "schedule_out": "15:00"},
    14: {"short_name": "Crystal", "full_name": "De Guzman, Crystal Joy A.",         "schedule_in": "06:00", "schedule_out": "15:00"},
    15: {"short_name": "Leigh",   "full_name": "Umali, Marie Anne Allaine L.",      "schedule_in": "06:00", "schedule_out": "15:00"},
    16: {"short_name": "Alli",    "full_name": "Mangana, Alliana Mhey M.",          "schedule_in": "06:00", "schedule_out": "15:00"},
    19: {"short_name": "Wela",    "full_name": "Coloma, Louela",                    "schedule_in": "06:00", "schedule_out": "18:00"},
    21: {"short_name": "Dyann",   "full_name": "Ching, Dyann Felicity Xyia A.",     "schedule_in": "06:00", "schedule_out": "15:00"},
    22: {"short_name": "Jhon",    "full_name": "Robledo, Jhon Albert C.",           "schedule_in": "06:00", "schedule_out": "15:00"},
    26: {"short_name": "Kath",    "full_name": "Mamalayan, Katherine Nicole",       "schedule_in": "09:00", "schedule_out": "15:00"},
}


# ── Default Holidays ──────────────────────────────────────────────────────────
DEFAULT_HOLIDAYS = [
    {"date": "2025-12-30", "type": "regular_holiday", "label": "December 30, 2025 - Rizal Day (Regular Holiday)",    "employee_id": "ALL", "value": 1},
    {"date": "2026-01-01", "type": "regular_holiday", "label": "January 1, 2026 - New Year's Day (Regular Holiday)", "employee_id": "ALL", "value": 1},
]


# ── Style Constants ───────────────────────────────────────────────────────────
THIN   = Side(style="thin",   color="000000")
MEDIUM = Side(style="medium", color="000000")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")

# ✅ ARIAL FONT CONSTANTS — define once, use everywhere
F_HEADER  = Font(name="Arial", bold=True, size=11)
F_NORMAL  = Font(name="Arial", size=11)
F_BOLD    = Font(name="Arial", size=11, bold=True)
F_TOTAL   = Font(name="Arial", size=11, bold=True)


# ── Dataclasses ───────────────────────────────────────────────────────────────
@dataclass
class RawPunch:
    employee_id: int
    punch_at: datetime
    department: str = ""
    verify_code: str = ""


@dataclass
class DailyRecord:
    employee_id: int
    display_name: str
    full_name: str
    work_date: date
    time_in: Optional[datetime]
    time_out: Optional[datetime]
    late_minutes: Optional[int]
    undertime_minutes: Optional[float]
    overtime_hours: Optional[int]
    half_day: Optional[float]
    remarks: Optional[str]


@dataclass
class Adjustment:
    date: Optional[date]
    employee_id: Optional[int]
    adj_type: str
    value: float
    label: str = ""
    remarks: str = ""


# ── Parsing Helpers ───────────────────────────────────────────────────────────
def parse_time_string(value: str) -> time:
    value = str(value).strip()
    for fmt in ("%H:%M", "%I:%M %p", "%I:%M:%S %p"):
        try:
            return datetime.strptime(value, fmt).time()
        except ValueError:
            continue
    raise ValueError(f"Invalid time value: {value}")


def parse_datetime_value(value: Any) -> Optional[datetime]:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    text = text.replace("a.m.", "am").replace("p.m.", "pm").replace("A.M.", "AM").replace("P.M.", "PM")
    for fmt in DATE_PATTERNS:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def fmt_dt(value: Optional[datetime]) -> Optional[str]:
    if value is None:
        return None
    hour = value.hour % 12 or 12
    ampm = value.strftime('%p').lower()
    return f"{value.day:02d}/{value.month:02d}/{value.year} {hour}:{value.minute:02d}:{value.second:02d} {ampm}"


def fmt_short(d: date) -> str:
    return f"{d.month}/{d.day}"


def excel_date_text(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


# ── Shift / Schedule Helpers ──────────────────────────────────────────────────
SHIFT_PRESETS: List[Tuple[str, str]] = [
    ("06:00", "15:00"),
    ("07:00", "16:00"),
    ("08:00", "17:00"),
    ("09:00", "18:00"),
    ("06:00", "18:00"),
]


def _safe_parse_time_or_none(value: Any) -> Optional[time]:
    text = str(value).strip() if value is not None else ""
    if not text:
        return None
    try:
        return parse_time_string(text)
    except Exception:
        return None


def _pick_shift_for_day(
    time_in: Optional[datetime],
    cfg_schedule_in: Optional[time],
    cfg_schedule_out: Optional[time],
) -> Tuple[time, time]:
    presets: List[Tuple[time, time]] = [
        (parse_time_string(start), parse_time_string(end))
        for start, end in SHIFT_PRESETS
    ]
    if cfg_schedule_in and cfg_schedule_out:
        cfg_pair = (cfg_schedule_in, cfg_schedule_out)
        if cfg_pair not in presets:
            presets.append(cfg_pair)

    if time_in is None:
        if cfg_schedule_in and cfg_schedule_out:
            return cfg_schedule_in, cfg_schedule_out
        return parse_time_string("08:00"), parse_time_string("17:00")

    punch_minutes = time_in.hour * 60 + time_in.minute + (time_in.second / 60.0)

    def score(pair: Tuple[time, time]) -> Tuple[float, float]:
        start_t, _ = pair
        start_minutes = start_t.hour * 60 + start_t.minute
        diff = abs(punch_minutes - start_minutes)
        return (diff, start_minutes)

    return min(presets, key=score)


# ── File Readers ──────────────────────────────────────────────────────────────
def _read_xlsx(file_bytes: bytes) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [
        str(ws.cell(1, c).value).strip() if ws.cell(1, c).value is not None else ""
        for c in range(1, ws.max_column + 1)
    ]
    rows: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        if any(v not in (None, "") for v in row.values()):
            rows.append(row)
    return rows


def _read_xls(file_bytes: bytes) -> List[Dict[str, Any]]:
    if xlrd is None:
        raise RuntimeError("xlrd is required to read .xls files. Run: pip install xlrd")
    book = xlrd.open_workbook(file_contents=file_bytes)
    sheet = book.sheet_by_index(0)
    headers = [str(sheet.cell_value(0, c)).strip() for c in range(sheet.ncols)]
    rows: List[Dict[str, Any]] = []
    for r in range(1, sheet.nrows):
        row: Dict[str, Any] = {}
        for c in range(sheet.ncols):
            value = sheet.cell_value(r, c)
            if isinstance(value, str):
                value = value.strip()
            row[headers[c]] = value
        if any(v not in (None, "") for v in row.values()):
            rows.append(row)
    return rows


def read_raw_rows(uploaded_file: Any) -> List[Dict[str, Any]]:
    if hasattr(uploaded_file, "getvalue"):
        name = getattr(uploaded_file, "name", "uploaded.xlsx")
        file_bytes = uploaded_file.getvalue()
    else:
        path = Path(uploaded_file)
        name = path.name
        file_bytes = path.read_bytes()
    lower = name.lower()
    if lower.endswith(".xlsx"):
        return _read_xlsx(file_bytes)
    if lower.endswith(".xls"):
        return _read_xls(file_bytes)
    raise ValueError("Only .xls and .xlsx files are supported.")


# ── Normalize Punches ─────────────────────────────────────────────────────────
def normalize_punches(raw_rows: List[Dict[str, Any]]) -> List[RawPunch]:
    punches: List[RawPunch] = []
    for row in raw_rows:
        emp_raw = (
            row.get("ID Number")
            or row.get("ID number")
            or row.get("ID")
            or row.get("ID No.")
            or row.get("No.")
            or row.get("No")
            or row.get("Employee ID")
            or row.get("employee_id")
            or row.get("Name")
        )
        dt_raw = (
            row.get("Date/Time")
            or row.get("Date Time")
            or row.get("Datetime")
            or row.get("Timestamp")
        )
        if emp_raw in (None, "") or dt_raw in (None, ""):
            continue
        try:
            employee_id = int(float(str(emp_raw).strip()))
        except ValueError:
            continue
        punch_at = parse_datetime_value(dt_raw)
        if punch_at is None:
            continue
        punches.append(
            RawPunch(
                employee_id=employee_id,
                punch_at=punch_at,
                department=str(row.get("Department") or "").strip(),
                verify_code=str(row.get("VerifyCode") or row.get("Verify Code") or "").strip(),
            )
        )
    punches.sort(key=lambda x: (x.employee_id, x.punch_at))
    return punches


# ── Adjustments ───────────────────────────────────────────────────────────────
def parse_adjustments_table(df_like_rows: Iterable[Dict[str, Any]]) -> List[Adjustment]:
    adjustments: List[Adjustment] = []
    for row in df_like_rows:
        if not row:
            continue
        adj_type = str(row.get("type") or row.get("Type") or "").strip().lower()
        if not adj_type:
            continue
        emp = row.get("employee_id") or row.get("Employee ID") or row.get("Employee")
        employee_id: Optional[int]
        if emp in (None, "", "ALL", "all"):
            employee_id = None
        else:
            try:
                employee_id = int(float(str(emp).strip()))
            except (ValueError, TypeError):
                employee_id = None
        dt_raw = row.get("date") or row.get("Date")
        parsed_dt = parse_datetime_value(dt_raw) if dt_raw not in (None, "") else None
        parsed_date = parsed_dt.date() if parsed_dt else None
        value = float(row.get("value") or row.get("Value") or 0)
        label = str(row.get("label") or row.get("Label") or "").strip()
        remarks = str(row.get("remarks") or row.get("Remarks") or "").strip()
        adjustments.append(Adjustment(parsed_date, employee_id, adj_type, value, label, remarks))
    return adjustments


def default_adjustments() -> List[Adjustment]:
    return [
        Adjustment(
            parse_datetime_value(item["date"]).date(),
            None if item["employee_id"] == "ALL" else int(item["employee_id"]),
            item["type"],
            float(item.get("value", 1)),
            item.get("label", ""),
            "",
        )
        for item in DEFAULT_HOLIDAYS
    ]


# ── Build Daily Records ───────────────────────────────────────────────────────
def build_daily_records(
    punches: List[RawPunch],
    employee_config: Dict[int, Dict[str, Any]],
    manual_adjustments: Optional[List[Adjustment]] = None,
) -> Tuple[List[DailyRecord], Dict[int, Dict[str, Any]], List[date]]:

    grouped: Dict[Tuple[int, date], List[datetime]] = {}
    for punch in punches:
        grouped.setdefault((punch.employee_id, punch.punch_at.date()), []).append(punch.punch_at)

    employee_meta: Dict[int, Dict[str, Any]] = {}
    for emp_id in sorted({p.employee_id for p in punches}):
        cfg = employee_config.get(emp_id, {})
        short_name = cfg.get("short_name") or str(emp_id)
        full_name  = cfg.get("full_name")  or short_name
        cfg_in  = _safe_parse_time_or_none(cfg.get("schedule_in", ""))
        cfg_out = _safe_parse_time_or_none(cfg.get("schedule_out", ""))
        employee_meta[emp_id] = {
            "short_name":   short_name,
            "full_name":    full_name,
            "schedule_in":  cfg_in  or parse_time_string("08:00"),
            "schedule_out": cfg_out or parse_time_string("17:00"),
        }

    adjustments_by_emp_date: Dict[Tuple[int, date], List[Adjustment]] = {}
    global_adjustments_by_date: Dict[date, List[Adjustment]] = {}

    if manual_adjustments:
        for adj in manual_adjustments:
            if adj.date is None:
                continue
            if adj.employee_id is None:
                global_adjustments_by_date.setdefault(adj.date, []).append(adj)
            else:
                adjustments_by_emp_date.setdefault((adj.employee_id, adj.date), []).append(adj)

    all_dates = sorted({work_date for (_, work_date) in grouped.keys()})
    records: List[DailyRecord] = []

    for (emp_id, work_date), events in sorted(grouped.items(), key=lambda item: (item[0][0], item[0][1])):
        meta = employee_meta[emp_id]
        events = sorted(events)

        time_in  = events[0]          if events           else None
        time_out = events[-1]         if len(events) >= 2 else None

        day_adjustments: List[Adjustment] = (
            adjustments_by_emp_date.get((emp_id, work_date), [])
            + global_adjustments_by_date.get(work_date, [])
        )
        adj_types = {adj.adj_type for adj in day_adjustments}

        if "force_no_time_in" in adj_types and events:
            time_in  = None
            time_out = events[-1]

        if "force_no_time_out" in adj_types and events:
            time_in  = events[0]
            time_out = None

        remarks: List[str] = []
        late:      Optional[int]   = None
        undertime: Optional[float] = None
        overtime:  Optional[int]   = None
        half_day:  Optional[float] = None

        shift_in, shift_out = _pick_shift_for_day(time_in, meta["schedule_in"], meta["schedule_out"])
        scheduled_in_dt  = datetime.combine(work_date, shift_in)
        scheduled_out_dt = datetime.combine(work_date, shift_out)

        if time_in and time_in > scheduled_in_dt:
            late_minutes = int((time_in - scheduled_in_dt).total_seconds() // 60)
            late = late_minutes if late_minutes > 0 else None

        if late is not None and late >= 120:
            half_day  = 0.5
            late      = None
            undertime = None
            overtime  = None
            if "half day" not in " ".join(remarks).lower():
                remarks.append("half day")

        if time_out and time_out < scheduled_out_dt:
            ut_mins = int((scheduled_out_dt - time_out).total_seconds() // 60)
            undertime = ut_mins if 0 < ut_mins <= 60 else None

        if time_out and time_out > scheduled_out_dt:
            ot_hours = int((time_out - scheduled_out_dt).total_seconds() / 3600)
            overtime = ot_hours if ot_hours > 0 else None

        if time_in is None:
            remarks.append("no time in")
        if time_out is None:
            remarks.append("no time out")

        for adj in day_adjustments:
            if adj.adj_type == "half_day":
                half_day = adj.value if adj.value else 0.5
                if "half day" not in " ".join(remarks).lower():
                    remarks.append("half day")
            elif adj.adj_type == "force_no_time_in":
                if "no time in" not in " ".join(remarks).lower():
                    remarks.append("no time in")
            elif adj.adj_type == "force_no_time_out":
                if "no time out" not in " ".join(remarks).lower():
                    remarks.append("no time out")
            if adj.remarks:
                remarks.append(adj.remarks)

        remarks_text_all = " ".join(remarks).lower()
        if half_day or "half day" in remarks_text_all:
            late      = None
            undertime = None
            overtime  = None

        seen = set()
        clean_remarks = []
        for item in remarks:
            text = item.strip()
            if text and text.lower() not in seen:
                seen.add(text.lower())
                clean_remarks.append(text)

        records.append(
            DailyRecord(
                employee_id=emp_id,
                display_name=meta["short_name"],
                full_name=meta["full_name"],
                work_date=work_date,
                time_in=time_in,
                time_out=time_out,
                late_minutes=late,
                undertime_minutes=undertime,
                overtime_hours=overtime,
                half_day=half_day,
                remarks=", ".join(clean_remarks) if clean_remarks else None,
            )
        )

    return records, employee_meta, all_dates


# ── Excel Helpers ─────────────────────────────────────────────────────────────
def _totals_row_formula(col: str, start_row: int, end_row: int) -> str:
    return f"=SUM({col}{start_row}:{col}{end_row})"


# ── Build Output Workbook ─────────────────────────────────────────────────────
def build_output_workbook(
    records: List[DailyRecord],
    employee_meta: Dict[int, Dict[str, Any]],
    cutoff_label: str = "",
    payroll_label: str = "",
    adjustments: Optional[List[Adjustment]] = None,
    prepared_by_name: str = "",
    prepared_by_title: str = "",
) -> Workbook:
    wb = Workbook()

    # ✅ Force Arial as default font for entire workbook
    wb._named_styles["Normal"].font = Font(name="Arial", size=11)

    ws = wb.active
    ws.title = "TIME IN & TIME OUT"
    build_time_in_out_sheet(ws, records)
    return wb


# ── Build Time In/Out Sheet ───────────────────────────────────────────────────
def build_time_in_out_sheet(ws, records: List[DailyRecord]) -> Dict[int, int]:
    widths = {
        "A": 4.71,  "B": 13.57, "C": 21.57, "D": 22.71,
        "E": 16.71, "F": 13.43, "G": 15.43, "H": 12.0, "I": 51.43,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    headers = [
        None, "Name", "Date/Time (IN)", "Date/Time (OUT)",
        "Late (Mins)", "Undertime", "Overtime (Hrs)", "Half Day", "Remarks",
    ]
    for idx, value in enumerate(headers, start=1):
        cell = ws.cell(1, idx, value)
        cell.font      = F_HEADER                          # ✅ Arial Bold 11
        cell.alignment = LEFT if idx == 2 else CENTER
        cell.border    = Border(bottom=MEDIUM)

    total_row_by_emp: Dict[int, int] = {}
    row = 3
    grouped: Dict[int, List[DailyRecord]] = {}
    for record in records:
        grouped.setdefault(record.employee_id, []).append(record)

    for emp_id in sorted(grouped.keys()):
        start_row = row
        for rec in grouped[emp_id]:
            ws.cell(row, 1, rec.employee_id).alignment      = CENTER
            ws.cell(row, 2, rec.display_name).alignment     = LEFT
            ws.cell(row, 3, fmt_dt(rec.time_in)).alignment  = CENTER
            ws.cell(row, 4, fmt_dt(rec.time_out)).alignment = CENTER
            ws.cell(row, 5, rec.late_minutes).alignment     = CENTER
            ws.cell(row, 6, rec.undertime_minutes).alignment = CENTER
            ws.cell(row, 7, rec.overtime_hours).alignment   = CENTER
            ws.cell(row, 8, rec.half_day).alignment         = CENTER
            ws.cell(row, 9, rec.remarks).alignment          = LEFT
            for c in range(1, 10):
                ws.cell(row, c).font = F_NORMAL             # ✅ Arial Regular 11
            row += 1

        end_row   = row - 1
        total_row = row
        total_row_by_emp[emp_id] = total_row

        ws.cell(total_row, 4, f"=COUNTA(B{start_row}:B{end_row})-SUM(H{start_row}:H{end_row})").alignment = CENTER
        ws.cell(total_row, 5, _totals_row_formula("E", start_row, end_row)).alignment = CENTER
        ws.cell(total_row, 6, _totals_row_formula("F", start_row, end_row)).alignment = CENTER
        ws.cell(total_row, 7, _totals_row_formula("G", start_row, end_row)).alignment = CENTER
        ws.cell(total_row, 8, _totals_row_formula("H", start_row, end_row)).alignment = CENTER

        for c in range(1, 10):
            ws.cell(total_row, c).font      = F_TOTAL       # ✅ Arial Bold 11
            ws.cell(total_row, c).border    = Border(top=THIN)
            ws.cell(total_row, c).alignment = LEFT if c in (2, 9) else CENTER
        row += 2

    return total_row_by_emp


# ── Save Workbook ─────────────────────────────────────────────────────────────
def save_workbook_to_bytes(wb: Workbook) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ── Label Helpers ─────────────────────────────────────────────────────────────
def format_date_label(d: date) -> str:
    return f"{d.strftime('%B')} {d.day}, {d.year}"


def infer_cutoff_label(records: List[DailyRecord]) -> str:
    if not records:
        return ""
    dates = sorted({r.work_date for r in records})
    return f"{format_date_label(dates[0])} - {format_date_label(dates[-1])}"


def infer_payroll_label(records: List[DailyRecord]) -> str:
    if not records:
        return ""
    dates = sorted({r.work_date for r in records})
    last_date = dates[-1]
    if last_date.day <= 15:
        payroll_date = date(last_date.year, last_date.month, 15)
    else:
        last_day = calendar.monthrange(last_date.year, last_date.month)[1]
        payroll_date = date(last_date.year, last_date.month, last_day)
    return format_date_label(payroll_date)